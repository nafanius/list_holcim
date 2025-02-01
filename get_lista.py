import threading
import ezsheets

import get_del_new_lists


import data_sql
import openpyxl
from datetime import time
import logging
from datetime import datetime, timedelta
import os
import re
from convert_lists import get_list_from_three_norm_del_add

# region logging
logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s"
)
lg = logging.debug
cr = logging.critical
inf = logging.info
exp = logging.exception
# logging.disable(logging.DEBUG)
# logging.disable(logging.INFO)
# logging.disable(logging.CRITICAL)
# endregion

db_lock = threading.Lock()

# region excel file
def generate_name_of_file_google():
    """Forms the file names that need to be downloaded based on the current date, returning a list of file names

    Returns:
        list of str: list of file names
    """
    list_of_download_files = []

    now = datetime.now()
    current_week_number = now.isocalendar()[1]
    current_year = now.year

    if current_week_number == 1 and current_year == (now - timedelta(weeks=1)).year:
        list_of_download_files.append(f"Tydz {1}.{current_year + 1}")
        list_of_download_files.append(f"Tydz {2}.{current_year + 1}")

    else:
        list_of_download_files.append(f"Tydz {current_week_number}.{current_year}")
        list_of_download_files.append(f"Tydz {current_week_number + 1}.{current_year}")

    return list_of_download_files


def save_google_sheet(directory="excel_files"):
    """save googlsheet file like a Excel in directory

    Args:
        directory (str, optional): The directory where the file is save. Defaults to 'excel_files'.
    """
    for file in generate_name_of_file_google():
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
            ss = ezsheets.Spreadsheet(file)
            ss_name = ss.title
            file_path = os.path.join(directory, f"{ss_name}.xlsx")
            ss.downloadAsExcel(file_path)
        except Exception as err:
            inf(err)
            inf("ERRR")
            continue


def form_lista_beton(excel_file, day, date_of_day_text):
    """We retrieve the shipping schedule from the excel_file and return lista_beton

    Args:
        excel_file (str): name of excel file
        day (int): number of the day  of the week
        date_of_day_text (str): date of day request - '12.03.2024'

    Returns:
        list: lista_beton, del_lista, add_lista: 3 list of cartage with old add del
    """
    lista_beton = []
    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as err:

        inf(f"{err}\nтакого файла нет " + excel_file)
        return []

    sheet = wb[wb.sheetnames[day]]

    def fill_list_beton(times_download, row, column, wenz):
        time_pattern = r"\b([01]?[0-9]|2[0-3]):([0-5][0-9])\b"
        match = re.search(time_pattern, str(times_download))
        if match:
            hours, minutes = match.groups()
            times_download = time(int(hours), int(minutes))
            lista_beton.append(
                (
                    sheet.cell(row=row, column=column + 4).value,
                    times_download,
                    sheet.cell(row=row, column=column + 1).value,
                    sheet.cell(row=row, column=column + 2).value,
                    sheet.cell(row=row, column=column + 10).value,
                    sheet.cell(row=row, column=column + 13).value,
                    sheet.cell(row=row, column=column + 11).value,
                    wenz,
                )
            )

    for row_in_file in range(1, sheet.max_row):
        find_line = sheet.cell(row=row_in_file, column=3).value
        if "zawodzie 2 " in str(find_line).lower():
            for row_in_beton in range(11, 42):
                fill_list_beton(
                    sheet.cell(row=row_in_file + row_in_beton, column=3).value,
                    row_in_file + row_in_beton,
                    3,
                    "2",
                )

        if "zawodzie 1 " in str(find_line).lower():
            for row_in_beton in range(11, 42):
                fill_list_beton(
                    sheet.cell(row=row_in_file + row_in_beton, column=3).value,
                    row_in_file + row_in_beton,
                    3,
                    "1",
                )

    lista_beton = sorted(lista_beton, key=lambda event: event[1])

   
    del_lista, add_lista = get_del_new_lists.check_del_add_lista(
        date_of_day_text, lista_beton
    )


    inf(del_lista)
    inf(add_lista)

    # добавил сохранение в базе данных листы бетона
    with db_lock:
        data_sql.record_beton({"date_of_day_text":date_of_day_text, "lista_beton":lista_beton, "day":day})
    
    return lista_beton, del_lista, add_lista

# endregion excel file

def form_lista(excel_file, day, date_of_day_text):
    """_summary_

    Args:
        excel_file (str): name of excel file
        day (int): the number of the day of the week

    Returns:
        list: Departure schedule list
    """
    lista = []
    try:
        wb = openpyxl.load_workbook(excel_file)
    except:
        inf("такого файла нет " + excel_file)
        return []
    sheet = wb[wb.sheetnames[day]]

    def fill_list(time_start, row, column):
        time_pattern = r"\b([01]?[0-9]|2[0-3]):([0-5][0-9])\b"
        match = re.search(time_pattern, str(time_start))
        if match:
            hours, minutes = match.groups()
            time_start = time(int(hours), int(minutes))
            lista.append(
                (time_start, str(sheet.cell(row=row, column=column - 1).value).strip())
            )

    for row_in_file in range(1, sheet.max_row):
        c = sheet.cell(row=row_in_file, column=3).value
        if "zawodzie 2 " in str(c).lower() or "zawodzie 1 " in str(c).lower():
            for row_in_list_time in range(15):
                fill_list(
                    sheet.cell(row=row_in_file + row_in_list_time, column=12).value,
                    row_in_file + row_in_list_time,
                    12,
                )
                fill_list(
                    sheet.cell(row=row_in_file + row_in_list_time, column=15).value,
                    row_in_file + row_in_list_time,
                    15,
                )

    lista = sorted(lista, key=lambda event: event[0])

    # добавил сохранение в базе данных lista
    with db_lock:
        data_sql.record_lista({"date_of_day_text":date_of_day_text, "lista":lista, "day":day})
    
    return lista


def lista_in_text(lista):
    """Generates a list in text form from the list of departure data

    Args:
        lista (list): List of date depurture shedule list

    Returns:
        list: The generated list for populating the departure list in HTML
    """

    if not lista:
        return []
    lista_text = []
    for time_in_list, person in lista:
        lista_text.append(f"{time_in_list.strftime('%H:%M')} {person}")

    return lista_text


def lista_in_text_beton(lista_beton):
    """Generates a list in text form from the list of shipment data

    Args:
        lista_beton (list): List of date depurture shipment list

    Returns:
        list: The generated list for populating the shipment list in HTML
    """

    def sum_of_metres(data, sort):
        sum_m = 0
        try:
            if sort in (0, 2):
                sum_m += float(data)
        except (ValueError, TypeError):
            # Игнорируем элементы, которые не являются числами
            pass
        return sum_m

    if not lista_beton:
        return ""
    
    lista_text = []
    sum_metres = 0

    for metres, times, firm, name, uwagi, przebieg, tel, wenz, sort in lista_beton:
        
        sum_metres = sum_metres + sum_of_metres(metres, sort)

        if sort == 0:
            lista_text += [
                f"{times} {metres} węzeł {wenz}",
                f"{firm}",
                f'{name} {uwagi + " " + przebieg}',
                f"{tel}",
                f"--------------------",
            ]
        elif sort == 1:
            lista_text += [
                f'<span style="color: rgb(238, 36, 36); font-weight: bold; text-decoration: line-through;">{times} {metres} węzeł {wenz}</span>',
                f'<span style="color: rgb(238, 36, 36); font-weight: bold; text-decoration: line-through;">{firm}</span>',
                f'<span style="color: rgb(238, 36, 36); font-weight: bold; text-decoration: line-through;">{name} {uwagi + " " + przebieg}</span>',
                f'<span style="color: rgb(238, 36, 36); font-weight: bold; text-decoration: line-through;">{tel}</span>',
                f'<span style="color: rgb(238, 36, 36); font-weight: bold; text-decoration: line-through;">--------------------</span>',
            ]
        elif sort == 2:
            lista_text += [
                f'<span style="color: rgb(0, 139, 7); font-weight: bold;">{times} {metres} węzeł {wenz}</span>',
                f'<span style="color: rgb(0, 139, 7); font-weight: bold;">{firm}</span>',
                f'<span style="color: rgb(0, 139, 7); font-weight: bold;">{name} {uwagi + " " + przebieg}</span>',
                f'<span style="color: rgb(0, 139, 7); font-weight: bold;">{tel}</span>',
                f'<span style="color: rgb(0, 139, 7); font-weight: bold;">--------------------</span>',
            ]

    return (
        lista_text,
        f'<p style="font-weight: bold; margin-bottom: 3px">zaplanowano metrów - {round(sum_metres, 1)}</p>',
    )


def find_day_request():
    """Generates a list of requests for three days depending on the current date
    and day of the week

    Returns:
        list: A list of three tuples containing the number of the day of the week,
              the name of the Excel file from which to take data, and the date string of
              the request day
    """
    list_of_days = []

    now = datetime.now()
    # Извлекаем номер недели с помощью isocalendar
    current_week_number = now.isocalendar()[1]
    day_of_week = now.weekday()
    current_year = now.year

    def create_week_and_year_to_file_name(type_of_days, week, year):
        if type_of_days == 1:
            if week == 1 and year == (now - timedelta(weeks=1)).year:
                return (1, year + 1), (1, year + 1), (1, year + 1)
            else:
                return (week, year), (week, year), (week, year)

        elif type_of_days == 2:
            if week == 1 and year == (now - timedelta(weeks=1)).year:
                return (1, year + 1), (1, year + 1), (2, year + 1)
            else:
                return (week, year), (week, year), (week + 1, year)

        elif type_of_days == 3:
            if week == 1 and year == (now - timedelta(weeks=1)).year:
                return (1, year + 1), (2, year + 1), (2, year + 1)
            else:
                return (week, year), (week + 1, year), (week + 1, year)

        elif type_of_days == 4:
            if week == 1 and year == (now - timedelta(weeks=1)).year:
                return (2, year + 1), (2, year + 1), (2, year + 1)
            else:
                return (week + 1, year), (week + 1, year), (week + 1, year)
        else:
            return []

    if day_of_week in (0, 1, 2, 3):
        weeks_years = create_week_and_year_to_file_name(
            1, current_week_number, current_year
        )

        list_of_days.append(
            (
                day_of_week,
                f"./excel_files/Tydz {weeks_years[0][0]}.{weeks_years[0][1]}.xlsx",
                now.strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                day_of_week + 1,
                f"./excel_files/Tydz {weeks_years[1][0]}.{weeks_years[1][1]}.xlsx",
                (now + timedelta(days=1)).strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                day_of_week + 2,
                f"./excel_files/Tydz {weeks_years[2][0]}.{weeks_years[2][1]}.xlsx",
                (now + timedelta(days=2)).strftime("%d.%m.%Y"),
            )
        )

    elif day_of_week == 4:
        weeks_years = create_week_and_year_to_file_name(
            2, current_week_number, current_year
        )
        list_of_days.append(
            (
                day_of_week,
                f"./excel_files/Tydz {weeks_years[0][0]}.{weeks_years[0][1]}.xlsx",
                now.strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                day_of_week + 1,
                f"./excel_files/Tydz {weeks_years[1][0]}.{weeks_years[1][1]}.xlsx",
                (now + timedelta(days=1)).strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                0,
                f"./excel_files/Tydz {weeks_years[2][0]}.{weeks_years[2][1]}.xlsx",
                (now + timedelta(days=3)).strftime("%d.%m.%Y"),
            )
        )
    elif day_of_week == 5:
        weeks_years = create_week_and_year_to_file_name(
            3, current_week_number, current_year
        )
        list_of_days.append(
            (
                day_of_week,
                f"./excel_files/Tydz {weeks_years[0][0]}.{weeks_years[0][1]}.xlsx",
                now.strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                0,
                f"./excel_files/Tydz {weeks_years[1][0]}.{weeks_years[1][1]}.xlsx",
                (now + timedelta(days=2)).strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                1,
                f"./excel_files/Tydz {weeks_years[2][0]}.{weeks_years[2][1]}.xlsx",
                (now + timedelta(days=3)).strftime("%d.%m.%Y"),
            )
        )
    elif day_of_week == 6:
        weeks_years = create_week_and_year_to_file_name(
            4, current_week_number, current_year
        )
        list_of_days.append(
            (
                0,
                f"./excel_files/Tydz {weeks_years[0][0]}.{weeks_years[0][1]}.xlsx",
                (now + timedelta(days=1)).strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                1,
                f"./excel_files/Tydz {weeks_years[1][0]}.{weeks_years[1][1]}.xlsx",
                (now + timedelta(days=2)).strftime("%d.%m.%Y"),
            )
        )
        list_of_days.append(
            (
                2,
                f"./excel_files/Tydz {weeks_years[2][0]}.{weeks_years[2][1]}.xlsx",
                (now + timedelta(days=3)).strftime("%d.%m.%Y"),
            )
        )

    return list_of_days


def combination_of_some_days_list():
    """Generates two dictionaries with three days of departure and shipment schedules

    Returns:
        two dict:  two dictionaries with three days of departure and shipment schedules
    """    
    day_of_week_list = [
        "poniedziałek",
        "wtorek",
        "środa",
        "czwartek",
        "piątek",
        "sobota",
        "niedziela",
    ]

    #  We save the latest Excel files from Google Drive.
    save_google_sheet()

    dict_list = {}
    dict_beton = {}
    list_of_day = find_day_request()
    
    number_elements = 1
    for day, file, date_of_day in list_of_day:
        split_text = lista_in_text(form_lista(file, day, date_of_day))
        if split_text == []:
            dict_list[f"element{number_elements}"] = [
                f"***{date_of_day} {day_of_week_list[day]}***",
                "Dane są niedostępne",
            ]
            number_elements += 1
            continue

        dict_list[f"element{number_elements}"] = [
            f"***{date_of_day} {day_of_week_list[day]}***",
            "",
        ] + split_text
        number_elements += 1

    for day, file, date_of_day in list_of_day:
        list_of_lists_norm_del_add = form_lista_beton(file, day, date_of_day)
        list_ready_to_covert_text = get_list_from_three_norm_del_add(
            *list_of_lists_norm_del_add
        )

        try:
            lista, meter = lista_in_text_beton(list_ready_to_covert_text)

            dict_beton[f"element{number_elements}"] = [
                f"***{date_of_day}  {day_of_week_list[day]}***",
                f"{meter}",
            ] + lista # type: ignore
            number_elements += 1
        except ValueError:
            dict_beton[f"element{number_elements}"] = [
                f"***{date_of_day} {day_of_week_list[day]}***",
                "Dane są niedostępne",
            ]
            number_elements += 1

    return dict_list | dict_beton


if __name__ == "__main__":
    inf(combination_of_some_days_list())
