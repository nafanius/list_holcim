import ezsheets

import openpyxl
from datetime import time
from datetime import datetime, timedelta
import os
import re


def generate_name_of_file_google():
    """генерируем файлы для загрузки"""
    list_of_download_files = []

    now = datetime.now()
    current_week_number = now.isocalendar()[1]
    current_year = now.year
    list_of_download_files.append(f"Tydz {current_week_number}.{current_year}")
    list_of_download_files.append(f"Tydz {current_week_number + 1}.{current_year}")

    return list_of_download_files


def get_from_google_sheet():
    for file in generate_name_of_file_google():
        try:
            directory = 'excel_files'
            ss = ezsheets.Spreadsheet(file)
            ss_name = ss.title
            file_path = os.path.join(directory, f'{ss_name}.xlsx')
            ss.downloadAsExcel(file_path)
        except:
            print("ERRR")
            continue


def form_lista(excel_file, day):
    '''дастоём расписание из файла'''
    lista_beton = []
    try:
        wb = openpyxl.load_workbook(excel_file)
    except:
        print("такого файла нет " + excel_file)
        return []

    sheet = wb[wb.sheetnames[day]]

    def fill_list_beton(times_download, row, column, wenz):
        if isinstance(times_download, time):
            lista_beton.append((sheet.cell(row=row, column=column + 4).value, times_download,
                                sheet.cell(row=row, column=column + 2).value,
                                sheet.cell(row=row, column=column + 10).value,
                                sheet.cell(row=row, column=column + 11).value, wenz))

    for row_in_file in range(1, sheet.max_row):
        c = sheet.cell(row=row_in_file, column=3).value
        if "zawodzie 2 " in str(c).lower():
            for row_in_beton in range(11, 42):
                fill_list_beton(sheet.cell(row=row_in_file + row_in_beton, column=3).value, row_in_file + row_in_beton,
                                3, "2")

        if "zawodzie 1 " in str(c).lower():
            for row_in_beton in range(11, 42):
                fill_list_beton(sheet.cell(row=row_in_file + row_in_beton, column=3).value, row_in_file + row_in_beton,
                                3, "1")

    lista_beton = sorted(lista_beton, key=lambda event: event[1])

    return lista_beton


def lista_in_bot(lista_beton):
    """"фотрмируем list в текстовый формат для высолки в бот """
    def sum_of_metres(data):
        sum_m = 0
        try:
            sum_m += float(data)
        except (ValueError, TypeError):
            # Игнорируем элементы, которые не являются числами
            pass
        return sum_m

    def convert_to_string(data):
        if not data:
            return ""
        try:
            data = str(data)
            data = data.strip()
            data = re.sub(r'\s+', ' ', data)
            return data
        except (TypeError, ValueError):
            return ""

    if not lista_beton:
        return ""
    lista_text = ""
    sum_metres = 0
    for metres, times, name, uwagi, tel, wenz in lista_beton:
        times = times.strftime('%H:%M')
        if tel:
            if isinstance(tel, float):
                tel = str(int(tel)).strip()
            elif isinstance(tel, str):
                tel = tel.strip()
        else:
            tel = ""

        name = convert_to_string(name)
        tel = convert_to_string(tel)
        uwagi = convert_to_string(uwagi)
        sum_metres = sum_metres + sum_of_metres(metres)
        metres = str(metres).strip()


        lista_text += (f"{times} {metres} węzeł {wenz}\n"
                       f"{name} {uwagi} {tel}\n"
                       f"--------------------\n")
    return lista_text, sum_metres


def find_day_request():
    list_of_days = []

    now = datetime.now()
    # Извлекаем номер недели с помощью isocalendar
    current_week_number = now.isocalendar()[1]
    day_of_week = now.weekday()
    current_year = now.year
    time_after_18_00 = now.replace(hour=18, minute=0, second=0, microsecond=0)
    if day_of_week in (0, 1, 2, 3, 4, 5):
        if now < time_after_18_00:
            list_of_days.append(
                (day_of_week, f"./excel_files/Tydz {current_week_number}.{current_year}.xlsx", now.strftime('%d.%m.%Y')))
        else:
            list_of_days.append((day_of_week + 1, f"./excel_files/Tydz {current_week_number}.{current_year}.xlsx",
                                 (now + timedelta(days=1)).strftime('%d.%m.%Y')))
    elif day_of_week == 6:
        list_of_days.append((0, f"./excel_files/Tydz {current_week_number + 1}.{current_year}.xlsx",
                             (now + timedelta(days=1)).strftime('%d.%m.%Y')))

    return list_of_days


def combination_of_some_days_list_bet():
    """формируем общий лист на несколько дней в зависимости от дня недели"""
    get_from_google_sheet()
    text_to_bot = ""
    for day, file, date in find_day_request():
        if not lista_in_bot(form_lista(file, day)):
            continue
        lista, meter = lista_in_bot(form_lista(file, day))
        text_to_bot += f"**{date}**\nMetres {meter}\n{lista}\n\n"

    return text_to_bot


if __name__ == '__main__':
    print(combination_of_some_days_list_bet())
